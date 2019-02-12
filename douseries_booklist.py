import time
import random
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook

#Some User Agents
hds=[
{'User-Agent':'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'},\
{'User-Agent':'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'},\
{'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'},\
{'User-Agent':'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US) AppleWebKit/534.16 (KHTML, like Gecko) Chrome/10.0.648.133 Safari/534.16'}
]

def book_spider(douseries_num):
	page_num = 1
	book_list = []
	try_times = 0
	
	while True:
		#https://book.douban.com/series/586?page=1
		douseries_url ='https://book.douban.com/series/'+str(douseries_num)+'?page='+str(page_num)
		
		time.sleep(random.random())
		
		try:
			req = requests.get(douseries_url,headers=hds[page_num%len(hds)])
	
			plain_txt = req.text
		except requests.exceptions.ConnectionError as e:
			print('获取网页url错误！————>',e)
			
		
		soup = BeautifulSoup(plain_txt)
		
		list_soup = soup.find_all('li',class_='subject-item')
		
		try_times +=1
		if list_soup == None and try_times<200:
			continue
		elif list_soup == None or len(list_soup)<=1:
			break
		
		for book_info in list_soup:
			#书名
			title = book_info.find_all('a')[1].stripped_strings #返回迭代器
			title_list =list(title)
			
			#链接
			book_link = book_info.find('a',class_='nbg').get('href')
			
			#作者，出版社，出版年
			desc = book_info.find('div',class_='pub').string.strip()
			desc_list = desc.split('/')
			
			
			#评分，评价人数
			rating_nums = book_info.find('span',class_='rating_nums')
			rating_pl = book_info.find('span',class_='pl')
			
			try:
				book_name = str(title_list[0]+title_list[1])
			except:
				book_name = str(title_list[0])
				
			try:
				star_info = '书籍评分： '+rating_nums.string.strip()
			except:
				star_info = '书籍评分： 暂无'
			try:
				pl_info = '评价人数： '+rating_pl.string.strip()
			except:
				pl_info = '评价人数：暂无'
					
			if len(desc_list) ==4:
				try:
					book_name = str(title_list[0]+title_list[1])
				except:
					book_name = str(title_list[0])
				
				try:
					author_info = '作者/译者：'+ desc_list[0].strip()
				except:
					author_info = '作者/译者：暂无'
					
				try:
					pub_info = '出版社： '+ desc_list[1].strip()
				except:
					pub_info = '出版社： 暂无'
					
				try:
					time_info = '出版年月： '+ desc_list[2].strip()
				except:
					time_info = '出版年月： 暂无'

				try:
					price_info = '价格： ' + desc_list[3].strip()
				except:
					price_info = '价格： 暂无'
				
			elif len(desc_list) ==5:
				try:
					author_info = '作者/译者：'+ desc_list[0].strip()+'/'+desc_list[1].strip()
				except:
					author_info = '作者/译者：暂无'
					
				try:
					pub_info = '出版社： '+  desc_list[2].strip()
				except:
					pub_info = '出版社： 暂无'
					
				try:
					time_info = '出版年月： '+ desc_list[3].strip()
				except:
					time_info = '出版年月： 暂无'

				try:
					price_info = '价格： ' + desc_list[4].strip()
				except:
					price_info = '价格： 暂无'
				
			
			book_list.append([book_name,author_info,pub_info,time_info,star_info,pl_info,book_link])
		
		print('获取书籍————第%d页'%page_num,'\n')
		page_num +=1
	
	douseries_title = soup.title.get_text()
	#douseries_about = soup.select('div[class=douseries_about')[0]
	
	print('\n总共收集到%d本书！'%len(book_list))
	return(book_list,douseries_title)
		

def do_spider(douseries_num):
	
	book_list,douseries_title= book_spider(douseries_num)
	
	book_list = sorted(book_list,key=lambda x:x[4],reverse=True)
	
	return(book_list,douseries_title)

def create_excel(path):	
	try:
		wordbook = load_workbook(path)
		print('\n已存在豆列书籍xlsx工作簿')
	except:
		print('\n未存在豆列数据xlsx工作簿，现在准备新建\n')
		wordbook = Workbook()
		wordbook.save(path)
		print('创建豆列书籍xlsx工作簿成功！')

	
def write_sheet(path,douseries_num,douseries_title,book_list):
	title = [['序号','书名', '作者/译者', '出版社', '出版年月', '书籍评分','评价人数','书籍链接'],]
	
	rows = len(title)
	
	wordbook = load_workbook(path)
	
	wsname = str(str(douseries_num)+str(douseries_title))
	
	sheet = wordbook.create_sheet(wsname,0)
	
	sheet.freeze_panes = 'A3'#A2 冻结第一行
	  
	#写入head_title，网页的标题
	sheet.cell(1,1,douseries_title)
	 
	#写入表头title
	for row1 in range(1,rows+1):
		for col1 in range(1,len(title[row1-1])+1):
			sheet.cell(row1+1,col1,title[row1-1][col1-1])#写入第二行
			
	print('\n创建豆列书籍xlsx工作表成功！')
			
	#写入书籍数据
	new_rows = len(book_list)
	
	xuhao = 1
	
	for row1 in range(1,new_rows+1):
		sheet.cell(row1+rows+1,1,xuhao)#从第三行
		xuhao +=1
		for col1 in range(1,len(book_list[row1-1])+1):
			sheet.cell(row1+rows+1,col1+1,book_list[row1-1][col1-1])#从第三行
	
	print('\n添加豆列书籍数据成功！')

	wordbook.save(path)
	


if __name__ =='__main__':
	
	#运行时间
	time.clock()
	
	#586 677 865 282
	douseries_num = 865
	
	book_save_path = r'F:\Python\Spider\doubanbooks\douseries_booklist(书籍丛书).xlsx'
	
	book_list,douseries_title= do_spider(douseries_num)
	
	create_excel(book_save_path)
	
	write_sheet(book_save_path,douseries_num,douseries_title,book_list)
	
	
	t1 = time.clock()
	print('\n 运行时间： ',t1)