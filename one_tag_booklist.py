import time
import requests
import numpy as np
from urllib import parse
from bs4 import BeautifulSoup
from imp import reload
from openpyxl import Workbook
from openpyxl import load_workbook

#Some User Agents
hds=[{'User-Agent':'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'},\
{'User-Agent':'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'},\
{'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'}]


#运行爬虫，获取某个标签下一页的全部书籍信息书籍信息
def book_spider(book_tag):
	#pages = page * 15
	page_nume = 0
	book_list = []
	try_times = 0
	
	while True:
		#设置打开豆瓣书籍标签页
		#http://www.douban.com/tag/%E5%B0%8F%E8%AF%B4/book?start=0
		#下一页增加15	
		tag_url='https://www.douban.com/tag/'+parse.quote(book_tag)+'/book?start='+str(page_nume)
		
		time.sleep(np.random.rand()*5)
	
		#
		try:
			#
			req = requests.get(tag_url,headers=hds[page_nume%len(hds)])
			#
			plain_txt = req.text
		except requests.exceptions.ConnectionError as e:
			print('获取网页url错误！————>',e)
		
		#
		soup = BeautifulSoup(plain_txt)
		#
		list_soup = soup.find('div',{'class':'mod book-list'})
		
		try_times +=1
		if list_soup == None and try_times<200:
			continue
		elif list_soup == None or len(list_soup)<=1:
			break

		for book_info in list_soup.find_all('dd'):
			#书名
			title = book_info.find('a',{'class':'title'}).string.strip()
			#以 / 分割的信息，作者，出版社，出版时间，价钱
			desc = book_info.find('div',{'class':'desc'}).string.strip()
			desc_list = desc.split('/')
			#书籍的链接
			book_url = book_info.find('a',{'class':'title'}).get('href')
			
			#书籍信息，评分
			try:
				author_info = '作者/译者: ' + '/'.join(desc_list[0:-3])
			except:
				author_info = '作者/译者: 暂无'			
			try:
				pub_info = '出版信息： '+'/'.join(desc_list[-3:])
			except:
				pub_info = '出版信息： 暂无'
			try:
				rating = '书籍评分： '+ book_info.find('span',{'class':'rating_nums'}).string.strip()
			except:
				rating = '书籍评分： 0.0'
			try:
				people_num = get_people_num(book_url)
				people_num = people_num.strip('人评价')
			except:
				people_num ='暂无'
				
			
			book_list.append([title,rating,people_num,author_info,pub_info])
		
		
		print('获取书籍————第%d页'%(page_nume/15+1),'\n')
		page_nume +=15
	
	print(book_list,'\n总共收集到%d本书！'%len(book_list))
	return(book_list)

	
def do_spider(book_tag):
	#全部书籍信息
	book_list = book_spider(book_tag)
	
	#按rating（评分）的降序排列
	book_list = sorted(book_list,key=lambda x:x[1],reverse=True)
	return(book_list)
	
	
def create_excel_xlsx(path,book_tag):
	title = [['序号','书名', '书籍评分', '评价人数', '作者/译者', '出版信息'],]
	
	rows = len(title)#获取需要写入数据的行数
	
	workbook = Workbook()#新建一个工作簿
	
	sheet = workbook.create_sheet(book_tag,0)#在工作簿开头，新建一个表格
	
	for row1 in range(1,rows+1):#从第1行开始
		for col1 in range(1,len(title[row1-1])+1):#从第1列开始
			sheet.cell(row1,col1,title[row1-1][col1-1])
	
			
	workbook.save(path)#保存工作簿
	print('xlsx格式表格写入数据成功！')
	

def write_excel_xlsx(path,book_tag,book_list):
	new_rows = len(book_list)#获取需要写入数据的新行数
	wordbook = load_workbook(path)#打开原工作簿
	#sheets = wordbook.sheetnames()#获取工作簿中的所有表格
	wordsheet = wordbook[book_tag]#获取特定的工作表
	old_rows = wordsheet.max_row #获取已存在的数据的行数

	xuhao = 1
	
	for i in range(1,new_rows+1):
		wordsheet.cell(i+old_rows,1,xuhao)
		xuhao +=1
		for j in range(1,len(book_list[i-1])+1):
			wordsheet.cell(i+old_rows,j+1,book_list[i-1][j-1])

	
	
	wordbook.save(path)
	print('xlsx格式表格【追加】数据成功！')

	

if __name__=='__main__':
	book_tag = '推理'
	#page = 3
	
	book_save_path = r'F:\Python\Spider\doubanbooks\\'+'one_tag_booklist--'+book_tag+'.xlsx'
	
	book_list = do_spider(book_tag)
	create_excel_xlsx(book_save_path,book_tag)
	write_excel_xlsx(book_save_path,book_tag,book_list)