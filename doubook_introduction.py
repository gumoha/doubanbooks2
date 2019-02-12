import time
import random
import requests
import re
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook


#Some User Agents
hds=hds=[
{'User-Agent':'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'},\
{'User-Agent':'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'},\
{'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'},\
{'User-Agent':'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US) AppleWebKit/534.16 (KHTML, like Gecko) Chrome/10.0.648.133 Safari/534.16'}
]

#打开xlsx文件，获取书籍链接
def collection_url(path):
	url_path = r'F:\Python\Spider\doubanbooks\豆瓣书籍链接.xlsx'
	
	url_list =[]
	wb = load_workbook(path)
	ws_list = wb.sheetnames[0:-1]#最后一张sheet1删除
	

	print('获取原xlsx文件的工作表成功！\n总共%s张'%len(ws_list),ws_list,'\n')

	#获取表格H行的url
	for ws in ws_list:
		ws1 = wb[ws]
		for row in ws1['H']:
			url_list.append(row.value)
	
	#删除表格表头 None，书籍链接
	while None in url_list:
		url_list.remove(None)
		while '书籍链接' in url_list:
			url_list.remove('书籍链接')

	url_nums = len(url_list)
	print('\nurl数量：%s'%url_nums,url_list)
	
	wb2 = Workbook()
	
	ws2 = wb2.create_sheet('url',0)
	
	for i in range(1,url_nums+1):
		ws2.cell(i,1,url_list[i-1])
	
	wb2.save(url_path)
	print('获取书籍链接,并保存xlsx成功')
	
	return url_list
	

def book_spider(url_list):
	book_list = []
	try_times = 0
	
	for url in url_list:
		time.sleep(random.random())
		
		try:
			req =requests.get(url,headers=hds[random.randrange(0,len(hds))])
			
			plain_txt = req.text
		except requests.exceptions.ConnectionError as e:
			print('获取网页url错误！————>',e)
		
		soup = BeautifulSoup(plain_txt)
		
		try_times +=1
		if soup == None and try_times<200:
			continue
		elif soup == None or len(soup)<=1:
			break
		
		
		#书名
		book_name = soup.find('a',class_='nbg').get('title')
		
		#获取书籍信息div
		info = soup.find('div',id='info')
		
		kongge = re.compile(r'[\s/]*')
		
		try:
			 ninfo_txt = re.sub(kongge,'',info.text)
		except:
			pass
	
		book_list.append([book_name,ninfo_txt,url])
	
	return book_list

def do_spider(url_list):
	
	book_list= book_spider(url_list)
	
	book_list = sorted(book_list,key=lambda x:x[0],reverse=True)
	
	print('\n 总共收集到 %s 本书！\n'%len(book_list))
	return(book_list)

	
def create_excel(save_path):	
	try:
		wordbook = load_workbook(save_path)
		print('\n已存在豆列书籍xlsx工作簿')
	except:
		print('\n未存在豆列数据xlsx工作簿，现在准备新建\n')
		wordbook = Workbook()
		wordbook.save(save_path)
		print('创建豆列书籍xlsx工作簿成功！')
	
def write_sheet(save_path,book_list):
	title = [['序号','书名', '书籍信息','书籍链接'],]
	
	rows = len(title)
	
	wordbook = load_workbook(save_path)
	
	sheet = wordbook.create_sheet('01',0)
	
	sheet.freeze_panes = 'A2'#A2 冻结第一行
	 
	#写入表头title
	for row1 in range(1,rows+1):
		for col1 in range(1,len(title[row1-1])+1):
			sheet.cell(row1,col1,title[row1-1][col1-1])#写入第一行
			
	print('\n创建豆列书籍xlsx工作表成功！')
			
	#写入书籍数据
	new_rows = len(book_list)
	
	xuhao = 1
	
	for row1 in range(1,new_rows+1):
		sheet.cell(row1+1,1,xuhao)#从第三行
		xuhao +=1
		for col1 in range(1,len(book_list[row1-1])+1):
			sheet.cell(row1+1,col1+1,book_list[row1-1][col1-1])#从第三行
	
	print('\n添加豆列书籍数据成功！')

	wordbook.save(save_path)
	

if __name__ == '__main__':
	#运行时间
	time.clock()
	
	collection_url_path = r'F:\Python\Spider\doubanbooks\test01.xlsx'
	
	save_path = r'F:\Python\Spider\doubanbooks\test02.xlsx'
	
	url_list = collection_url(collection_url_path)
	
	book_list = do_spider(url_list)
		
	create_excel(save_path)

	write_sheet(save_path,book_list)

	t1 = time.clock()
	print('\n 运行时间： ',t1)
	
	
	
	
	
	
	